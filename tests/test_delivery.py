"""Delivery（受注後・納品）アサイン計画のDB層テスト（#75）。

一時DBのみ使用。本番DB(cowork_sfa.db)には一切触れない。
検証: 提案到達での自動起票 / 見込み(提案)・確定(受注)の振り分け /
      クローズ済み非受注の除外 / アサインブロックの週展開・合算 / ベース工数合算。
"""
from __future__ import annotations

import re
import shutil
import tempfile
from pathlib import Path

import pytest

from cowork import sfa_db, webapp


@pytest.fixture
def con():
    d = tempfile.mkdtemp(prefix="sfa_delivery_test_")
    db = str(Path(d) / "t.db")
    sfa_db.init_db(db)
    conn = sfa_db.connect(db)
    try:
        yield conn
    finally:
        conn.close()
        shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def acc_id(con):
    aid = con.execute("INSERT INTO accounts(name) VALUES('テスト社')").lastrowid
    con.commit()
    return aid


def _deal(con, acc_id, stage, status="open", name="D"):
    return sfa_db.upsert_deal(con, account_id=acc_id, deal_name=name, stage=stage, status=status)


# ---- 自動起票（提案到達以降） ----

def test_ensure_delivery_only_from_proposal(con, acc_id):
    d_req = _deal(con, acc_id, "要件詰め")
    d_prop = _deal(con, acc_id, "提案")
    assert sfa_db.ensure_delivery_on_stage(con, d_req, "要件詰め") is None      # 提案未満は起票しない
    assert sfa_db.ensure_delivery_on_stage(con, d_prop, "提案") is not None     # 提案で起票
    # 二重起票しない
    assert sfa_db.ensure_delivery_on_stage(con, d_prop, "提案") is None
    assert len(sfa_db.list_deliveries(con, deal_id=d_prop)) == 1
    assert len(sfa_db.list_deliveries(con, deal_id=d_req)) == 0


def test_deal_form_shows_delivery_button_in_top_row_when_triggered(con, acc_id):
    """商談編集画面の上部ボタン列（＋新規開発案件等と並び）にも「＋Delivery追加」を出す。
    下部カードと同条件（既存Deliveryあり or 提案以降のステージ）のみ表示し、
    下までスクロールしないと気付けないという指摘への対応。"""
    d_won = _deal(con, acc_id, "受注", name="調達BPO")
    html = webapp.deal_form(con, sfa_db.get_deal(con, d_won))
    assert "🚚 ＋Delivery追加" in html
    assert html.find("🚚 ＋Delivery追加") < html.find("この商談を複製")

    d_early = _deal(con, acc_id, "初回アポ実施", name="早期商談")
    html_early = webapp.deal_form(con, sfa_db.get_deal(con, d_early))
    assert "🚚 ＋Delivery追加" not in html_early


def test_create_delivery_accepts_confidence_override_at_creation(con, acc_id):
    """新規Delivery起票時に確度（確定/見込み等）を指定できる（ユーザー要望2026-08-23）。
    不正値・省略時は自動導出(None)にフォールバックする。"""
    d = _deal(con, acc_id, "提案")
    dvid = sfa_db.create_delivery(con, deal_id=d, title="X", confidence_override="確定")
    dv = sfa_db.get_delivery(con, dvid)
    assert dv["confidence_override"] == "確定"
    assert sfa_db.delivery_confidence_effective(dv) == "確定"

    dvid2 = sfa_db.create_delivery(con, deal_id=d, title="Y", confidence_override="でたらめ")
    assert sfa_db.get_delivery(con, dvid2)["confidence_override"] is None

    dvid3 = sfa_db.create_delivery(con, deal_id=d, title="Z")
    assert sfa_db.get_delivery(con, dvid3)["confidence_override"] is None


def test_deal_form_and_deliveries_page_creation_forms_include_confidence_select(con, acc_id):
    d = _deal(con, acc_id, "受注", name="調達BPO")
    html = webapp.deal_form(con, sfa_db.get_deal(con, d))
    assert html.count('name="confidence_override"') == 2  # 上部ボタン列＋下部カードの両方

    dp_html = webapp.deliveries_page(con)
    assert 'name="confidence_override"' in dp_html


def test_delivery_title_defaults_to_deal_name(con, acc_id):
    did = _deal(con, acc_id, "受注", name="納品対象案件")
    sfa_db.ensure_delivery_on_stage(con, did, "受注")
    dv = sfa_db.list_deliveries(con, deal_id=did)[0]
    assert dv["title"] == "納品対象案件"


# ---- アサイン集計（見込み/確定・除外・週展開） ----

def test_compute_load_forecast_committed_and_exclusion(con, acc_id):
    d_prop = _deal(con, acc_id, "提案", name="見込み(提案中)案件")
    d_closing = _deal(con, acc_id, "クロージング", name="見込み(クロージング)案件")
    d_won = _deal(con, acc_id, "受注", name="確定案件")
    d_lost = _deal(con, acc_id, "提案", status="closed", name="失注案件")
    for did in (d_prop, d_closing, d_won, d_lost):
        sfa_db.ensure_delivery_on_stage(con, did, sfa_db.get_deal(con, did)["stage"])
    dv = {d["deal_id"]: d["id"] for d in sfa_db.list_deliveries(con)}
    W0 = "2026-07-27"  # 月曜
    sfa_db.add_delivery_assignment(con, delivery_id=dv[d_prop], owner="早瀬",
                                   from_week=W0, to_week="2026-08-10", fte_pct=50)
    sfa_db.add_delivery_assignment(con, delivery_id=dv[d_closing], owner="早瀬",
                                   from_week=W0, to_week="2026-08-10", fte_pct=30)
    sfa_db.add_delivery_assignment(con, delivery_id=dv[d_won], owner="早瀬",
                                   from_week=W0, to_week="2026-08-03", fte_pct=80)
    sfa_db.add_delivery_assignment(con, delivery_id=dv[d_lost], owner="早瀬",
                                   from_week=W0, to_week="2026-08-03", fte_pct=100)
    load = sfa_db.compute_delivery_load(con, start_week=W0, n_weeks=4)
    cell = load["cells"]["早瀬"][W0]
    assert cell["actual"]["proposal"] == 50    # 提案案件のみ（失注案件は除外）
    assert cell["actual"]["closing"] == 30     # クロージング案件
    assert cell["actual"]["committed"] == 80   # 受注案件
    # billingはfte_billing未指定なら実想定と同値
    assert cell["billing"]["proposal"] == 50
    # 範囲外の週は0（案件は8/10まで/8/3まで）
    assert "2026-08-17" not in load["cells"]["早瀬"]


def test_confidence_auto_derivation():
    assert sfa_db.delivery_confidence_auto("提案", "open") == "見込み(提案中)"
    assert sfa_db.delivery_confidence_auto("クロージング", "open") == "見込み(クロージング)"
    assert sfa_db.delivery_confidence_auto("受注", "open") == "確定"
    assert sfa_db.delivery_confidence_auto("受注", "closed") == "確定"   # 受注クローズは確定のまま
    assert sfa_db.delivery_confidence_auto("提案", "closed") == "無効(終了)"   # 失注等


def test_confidence_override_takes_priority_over_auto(con, acc_id):
    did = _deal(con, acc_id, "提案")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="D")
    dv = sfa_db.get_delivery(con, dv_id)
    assert sfa_db.delivery_confidence_effective(dv) == "見込み(提案中)"   # override無し=自動
    sfa_db.update_delivery(con, dv_id, confidence_override="確定")
    dv = sfa_db.get_delivery(con, dv_id)
    assert sfa_db.delivery_confidence_effective(dv) == "確定"   # 手修正が自動導出を上書き
    sfa_db.update_delivery(con, dv_id, confidence_override=None)
    dv = sfa_db.get_delivery(con, dv_id)
    assert sfa_db.delivery_confidence_effective(dv) == "見込み(提案中)"   # 空に戻すと自動に戻る


def test_delivery_is_active_excludes_completed_hold_cancelled_and_invalid(con, acc_id):
    did = _deal(con, acc_id, "提案")
    dv_ing = sfa_db.get_delivery(con, sfa_db.create_delivery(con, deal_id=did, status="進行中"))
    dv_done = sfa_db.get_delivery(con, sfa_db.create_delivery(con, deal_id=did, status="完了"))
    dv_hold = sfa_db.get_delivery(con, sfa_db.create_delivery(con, deal_id=did, status="保留"))
    dv_stop = sfa_db.get_delivery(con, sfa_db.create_delivery(con, deal_id=did, status="中止"))
    assert sfa_db.delivery_is_active(dv_ing) is True
    assert sfa_db.delivery_is_active(dv_done) is False
    assert sfa_db.delivery_is_active(dv_hold) is False
    assert sfa_db.delivery_is_active(dv_stop) is False
    # 状態=進行中でも、確度が(手修正で)無効(終了)ならactive対象外
    _id = sfa_db.create_delivery(con, deal_id=did, status="進行中")
    sfa_db.update_delivery(con, _id, confidence_override="無効(終了)")
    dv_invalid = sfa_db.get_delivery(con, _id)
    assert sfa_db.delivery_is_active(dv_invalid) is False


def test_compute_load_excludes_completed_hold_cancelled_deliveries(con, acc_id):
    """状態=完了/保留/中止のDeliveryは、紐づく商談が開いていても稼働集計から除外される。"""
    d_done = _deal(con, acc_id, "提案", name="完了済み案件")
    d_hold = _deal(con, acc_id, "提案", name="保留中案件")
    d_stop = _deal(con, acc_id, "提案", name="中止案件")
    d_active = _deal(con, acc_id, "提案", name="進行中案件")
    dv_done = sfa_db.create_delivery(con, deal_id=d_done, status="完了")
    dv_hold = sfa_db.create_delivery(con, deal_id=d_hold, status="保留")
    dv_stop = sfa_db.create_delivery(con, deal_id=d_stop, status="中止")
    dv_active = sfa_db.create_delivery(con, deal_id=d_active, status="進行中")
    W0 = "2026-07-27"
    for dv_id in (dv_done, dv_hold, dv_stop, dv_active):
        sfa_db.add_delivery_assignment(con, delivery_id=dv_id, owner="早瀬",
                                       from_week=W0, to_week="2026-08-03", fte_pct=40)
    load = sfa_db.compute_delivery_load(con, start_week=W0, n_weeks=2)
    cell = load["cells"]["早瀬"][W0]
    assert cell["actual"]["proposal"] == 40   # 進行中案件のみ
    assert len(load["items"]) == 1
    assert load["items"][0]["delivery_id"] == dv_active


def test_deliveries_page_sort_order(con, acc_id):
    """一覧の並び順: 確定→見込み(クロージング)→見込み(提案中)（進行中のみ・開始週の早い順）
    →保留→完了→中止→無効(終了)（確度=無効(終了)は状態に関わらず最下位）。"""
    def mkdeal(stage, status="open"):
        return sfa_db.upsert_deal(con, account_id=acc_id, deal_name="D", stage=stage, status=status)

    d_prop_late = mkdeal("提案")
    sfa_db.create_delivery(con, deal_id=d_prop_late, title="提案-遅", start_week="2026-09-01")
    d_prop_early = mkdeal("提案")
    sfa_db.create_delivery(con, deal_id=d_prop_early, title="提案-早", start_week="2026-08-01")
    d_closing = mkdeal("クロージング")
    sfa_db.create_delivery(con, deal_id=d_closing, title="クロージング", start_week="2026-08-15")
    d_won = mkdeal("受注")
    sfa_db.create_delivery(con, deal_id=d_won, title="確定", start_week="2026-08-20")
    d_hold = mkdeal("提案")
    sfa_db.create_delivery(con, deal_id=d_hold, title="保留", status="保留", start_week="2026-07-01")
    d_done = mkdeal("提案")
    sfa_db.create_delivery(con, deal_id=d_done, title="完了", status="完了", start_week="2026-07-01")
    d_stop = mkdeal("提案")
    sfa_db.create_delivery(con, deal_id=d_stop, title="中止", status="中止", start_week="2026-07-01")
    d_lost = mkdeal("提案")
    sfa_db.create_delivery(con, deal_id=d_lost, title="無効", start_week="2026-06-01")
    sfa_db.close_deal_to_lead(con, d_lost, "失注")   # status=中止になるが確度=無効(終了)が優先で最下位

    html = webapp.deliveries_page(con)
    order = re.findall(r'<input type="text" value="([^"]*)"', html)
    assert order == ["確定", "クロージング", "提案-早", "提案-遅", "保留", "完了", "中止", "無効"]


def test_base_workload_sum_and_upsert(con):
    sfa_db.upsert_base_workload(con, "早瀬", "営業", 30)
    sfa_db.upsert_base_workload(con, "早瀬", "管理", 10)
    assert sfa_db.base_workload_by_owner(con)["早瀬"] == 40
    # 同一(owner,function)はupsertで上書き（重複行にしない）
    sfa_db.upsert_base_workload(con, "早瀬", "営業", 25)
    assert sfa_db.base_workload_by_owner(con)["早瀬"] == 35
    assert len(sfa_db.list_base_workload(con, owner="早瀬")) == 2


def test_delivery_grid_expansion(con, acc_id):
    did = _deal(con, acc_id, "受注")
    sfa_db.ensure_delivery_on_stage(con, did, "受注")
    dvid = sfa_db.list_deliveries(con, deal_id=did)[0]["id"]
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="中島",
                                   from_week="2026-07-27", to_week="2026-08-10", fte_pct=60)
    grid = sfa_db.delivery_grid(con, dvid)
    assert grid["weeks"] == ["2026-07-27", "2026-08-03", "2026-08-10"]
    assert grid["cells"]["中島"]["2026-08-03"]["actual"] == 60


def test_billing_and_update(con, acc_id):
    did = _deal(con, acc_id, "受注")
    sfa_db.ensure_delivery_on_stage(con, did, "受注")
    dvid = sfa_db.list_deliveries(con, deal_id=did)[0]["id"]
    aid = sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="早瀬",
                                         from_week="2026-07-27", to_week="2026-07-27",
                                         fte_pct=80, fte_billing=100)
    load = sfa_db.compute_delivery_load(con, start_week="2026-07-27", n_weeks=1)
    c = load["cells"]["早瀬"]["2026-07-27"]
    assert c["actual"]["committed"] == 80 and c["billing"]["committed"] == 100
    # 編集: 実想定/請求/メンバーを更新
    sfa_db.update_delivery_assignment(con, aid, owner="中島", from_week="2026-07-27",
                                      to_week="2026-07-27", fte_pct=40, fte_billing=60, note="改")
    b = sfa_db.list_delivery_assignments(con, dvid)[0]
    assert b["owner"] == "中島" and b["fte_pct"] == 40 and b["fte_billing"] == 60 and b["note"] == "改"


def test_delivery_cascade_delete(con, acc_id):
    did = _deal(con, acc_id, "受注")
    sfa_db.ensure_delivery_on_stage(con, did, "受注")
    dvid = sfa_db.list_deliveries(con, deal_id=did)[0]["id"]
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="早瀬",
                                   from_week="2026-07-27", to_week="2026-07-27", fte_pct=10)
    sfa_db.delete_delivery(con, dvid)
    assert sfa_db.list_deliveries(con, deal_id=did) == []
    assert sfa_db.list_delivery_assignments(con, dvid) == []


# ---- 報酬額（月額↔総額 換算）＋ xlsx出力（#75 追加） ----

def test_delivery_month_count():
    # 月数＝合計週数÷4（4週≒1ヶ月）で統一
    assert sfa_db.delivery_month_count("2026-07-10", "2026-09-19") == 2.75   # 11週
    assert sfa_db.delivery_month_count("2026-07-06", "2026-07-27") == 1.0    # 4週
    assert sfa_db.delivery_month_count("2026-01-01", "2026-12-31") == 13.25  # 53週
    assert sfa_db.delivery_month_count("2026-09-28", "2026-11-16") == 2.0    # 8週
    assert sfa_db.delivery_month_count(None, None) == 1.0  # 未設定は1


def test_delivery_fee_monthly_to_total():
    mo, to = sfa_db.compute_delivery_fee("monthly", 100, None, 3)
    assert (mo, to) == (100, 300.0)


def test_delivery_fee_total_to_monthly():
    mo, to = sfa_db.compute_delivery_fee("total", None, 300, 3)
    assert (mo, to) == (100.0, 300)
    # 空入力は None
    assert sfa_db.compute_delivery_fee("monthly", "", "", 3) == (None, None)


def test_delivery_fee_both_present_is_trusted_as_is_manual_override():
    """新規タスク: 両方に値がある場合は再計算せずそのまま尊重する（自動換算後の手修正を保存で
    上書きしないため）。mode='monthly'でも、総額側に自動換算値と異なる手修正値が入っていれば
    その値のまま返る。"""
    # 通常の自動換算のまま(整合済み)なら当然そのまま
    mo, to = sfa_db.compute_delivery_fee("monthly", 100, 300, 3)
    assert (mo, to) == (100, 300)
    # 手修正: 月額100・月数3なら本来総額300のはずだが、人間が総額を850に手修正した場合
    mo, to = sfa_db.compute_delivery_fee("monthly", 100, 850, 3)
    assert (mo, to) == (100, 850), "手修正した総額が保存時に自動換算で上書きされてはいけない"
    # 逆方向(mode='total'でも同様に、月額側の手修正が尊重される)
    mo, to = sfa_db.compute_delivery_fee("total", 999, 300, 3)
    assert (mo, to) == (999, 300)


def test_delivery_form_fee_fields_allow_manual_edit_via_shared_js(con):
    """新規タスク: 灰色側(自動算出側)がreadOnlyでなくなり、手修正用のJSに委譲していること。"""
    from cowork import webapp
    acc = con.execute("INSERT INTO accounts(name) VALUES('A社')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    con.commit()
    dvid = sfa_db.create_delivery(con, deal_id=did, title="納品X",
                                  start_week="2026-07-06", end_week="2026-09-14")
    con.commit()
    html = webapp.delivery_form(con, dvid)
    assert "readOnly=true" not in html and "readOnly=false" not in html
    assert 'oninput="dvFeeFieldInput(this)"' in html
    assert 'onchange="dvFeeModeChanged()"' in html
    assert "function dvFeeFieldInput(" in html and "function dvFeeModeChanged(" in html


def test_delivery_fee_persist_and_xlsx(con):
    import io
    import openpyxl
    from cowork import webapp
    acc = con.execute("INSERT INTO accounts(name) VALUES('A社')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    con.commit()
    dvid = sfa_db.create_delivery(con, deal_id=did, title="納品X",
                                  start_week="2026-07-06", end_week="2026-09-14")
    con.commit()
    months = sfa_db.delivery_month_count("2026-07-06", "2026-09-14")  # 11週÷4 = 2.75
    assert months == 2.75
    mo, to = sfa_db.compute_delivery_fee("total", None, 300, months)
    sfa_db.update_delivery(con, dvid, fee_mode="total", fee_monthly=mo, fee_total=to)
    con.commit()
    r = sfa_db.get_delivery(con, dvid)
    assert r["fee_mode"] == "total" and r["fee_total"] == 300 and r["fee_monthly"] == 109.09
    # xlsx出力（全テーブル情報が3シートで出る）
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="早瀬",
                                   from_week="2026-07-06", to_week="2026-09-14",
                                   fte_pct=80, fte_billing=100, role="PM", member_kind="内部")
    con.commit()
    xls = webapp.build_deliveries_xlsx(con)
    assert xls[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    assert wb.sheetnames == ["Delivery一覧", "アサイン明細", "体制(役割別目標)", "月別入金計画"]
    ws = wb["Delivery一覧"]
    assert ws.cell(2, 9).value == 2.75         # 月数（11週÷4）
    assert ws.cell(2, 12).value == "総額報酬"   # 報酬形態（事業種別L1/L2列が2列追加され列位置が後ろへ移動）
    assert wb["アサイン明細"].cell(2, 7).value == "早瀬"


def test_delivery_total_assign_effort(con):
    acc = con.execute("INSERT INTO accounts(name) VALUES('A社')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="X", stage="受注", status="open")
    con.commit()
    dvid = sfa_db.create_delivery(con, deal_id=did, title="X",
                                  start_week="2026-05-18", end_week="2026-07-27")
    con.commit()
    # 5/18〜7/27 = 11週
    assert sfa_db._assignment_weeks("2026-05-18", "2026-07-27") == 11
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="高橋",
                                   from_week="2026-05-18", to_week="2026-07-27",
                                   fte_pct=20, fte_billing=40, role="リード")
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="杉山",
                                   from_week="2026-05-18", to_week="2026-07-27",
                                   fte_pct=80, fte_billing=80, role="コンサル")
    con.commit()
    # 期間平均の合計稼働率: (11*20 + 11*80) / 11週 = 1100/11 = 100.0（%/月）
    assert sfa_db.delivery_total_assign_effort(con, dvid) == 100.0
    # 一部期間のみのアサインは期間割りで薄まる（杉山を後半6週=6/22〜7/27だけ80%に変更）
    sugi = [a for a in sfa_db.list_delivery_assignments(con, dvid) if a["owner"] == "杉山"][0]
    sfa_db.update_delivery_assignment(con, sugi["id"], owner="杉山",
                                      from_week="2026-06-22", to_week="2026-07-27",
                                      fte_pct=80, fte_billing=80)
    con.commit()
    # 高橋20%×11週=220, 杉山80%×6週=480 → 700/11 ≈ 63.6
    assert sfa_db.delivery_total_assign_effort(con, dvid) == 63.6


def test_delivery_assign_effort_billing_basis(con):
    """平均単価用: 請求ベース工数（請求>0は請求、請求0/未入力は実想定にフォールバック）。"""
    acc = con.execute("INSERT INTO accounts(name) VALUES('X社')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="D", stage="受注", status="open")
    con.commit()
    dvid = sfa_db.create_delivery(con, deal_id=did, title="D",
                                  start_week="2026-05-18", end_week="2026-07-27")
    con.commit()
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="高橋",
                                   from_week="2026-05-18", to_week="2026-07-27",
                                   fte_pct=20, fte_billing=40)
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="杉山",
                                   from_week="2026-05-18", to_week="2026-07-27",
                                   fte_pct=80, fte_billing=80)
    con.commit()
    # 実想定: (20+80)=100 / 請求: (40+80)=120
    assert sfa_db.delivery_total_assign_effort(con, dvid) == 100.0
    assert sfa_db.delivery_total_assign_effort(con, dvid, use_billing=True) == 120.0
    # 高橋の請求を0にすると純粋請求は0扱い（フォールバックなし）→ 請求ベースは杉山80のみ = 80
    taka = [a for a in sfa_db.list_delivery_assignments(con, dvid) if a["owner"] == "高橋"][0]
    sfa_db.update_delivery_assignment(con, taka["id"], owner="高橋",
                                      from_week="2026-05-18", to_week="2026-07-27",
                                      fte_pct=20, fte_billing=0)
    con.commit()
    assert sfa_db.delivery_total_assign_effort(con, dvid, use_billing=True) == 80.0


def test_delivery_grid_ignores_blank_week_rows(con):
    """開始/終了週が空のアサイン行（役割追加直後など）があっても delivery_grid が落ちない（#502回避）。"""
    acc = con.execute("INSERT INTO accounts(name) VALUES('浜松')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="X", stage="受注", status="open")
    con.commit()
    dvid = sfa_db.create_delivery(con, deal_id=did, title="X")
    con.commit()
    # 空週の行を直挿し（from_week NOT NULL のため空文字）
    con.execute("INSERT INTO delivery_assignments(delivery_id,role,member_kind,owner,from_week,to_week,fte_pct) "
                "VALUES(?,?,?,?,?,?,?)", (dvid, "リード", "内部", "", "", "", 0))
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="杉山",
                                   from_week="2026-05-18", to_week="2026-07-27", fte_pct=80)
    con.commit()
    g = sfa_db.delivery_grid(con, dvid)          # 例外を投げないこと
    assert g["owners"] == ["杉山"] and len(g["weeks"]) == 11
    # 空週の行しかない場合は空グリッド（例外なし）
    dvid2 = sfa_db.create_delivery(con, deal_id=did, title="Y")
    con.commit()
    con.execute("INSERT INTO delivery_assignments(delivery_id,role,member_kind,owner,from_week,to_week,fte_pct) "
                "VALUES(?,?,?,?,?,?,?)", (dvid2, "リード", "内部", "", "", "", 0))
    con.commit()
    assert sfa_db.delivery_grid(con, dvid2) == {"weeks": [], "owners": [], "cells": {}}


def test_delivery_unit_price_8weeks_case(con):
    """8週・チーム合計100%・総額300万 → 月数=2(=8週/4)・月額150万・平均単価(月額)150万。"""
    import io
    import openpyxl
    from cowork import webapp
    acc = con.execute("INSERT INTO accounts(name) VALUES('マルハン')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="AI受託", stage="クロージング", status="open")
    con.commit()
    dvid = sfa_db.create_delivery(con, deal_id=did, title="AI受託",
                                  start_week="2026-09-28", end_week="2026-11-16")  # 8週
    con.commit()
    months = sfa_db.delivery_month_count("2026-09-28", "2026-11-16")
    assert months == 2.0
    mo, to = sfa_db.compute_delivery_fee("total", None, 300, months)
    assert (mo, to) == (150.0, 300.0)   # 総額300 ÷ 2ヶ月 = 月額150
    sfa_db.update_delivery(con, dvid, fee_mode="total", fee_monthly=mo, fee_total=to)
    # チーム合計100%（実想定30+50+20、請求は0＝実想定にフォールバック）
    for ow, pct in (("早瀬", 30), ("杉山", 50), ("戸野", 20)):
        sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner=ow,
                                       from_week="2026-09-28", to_week="2026-11-16",
                                       fte_pct=pct, fte_billing=0)
    con.commit()
    # 実想定は100%/月。請求は全行0%＝純粋請求は0（成果物ベース＝稼働コミットなし）
    assert sfa_db.delivery_total_assign_effort(con, dvid) == 100.0
    assert sfa_db.delivery_total_assign_effort(con, dvid, use_billing=True) == 0.0
    # 平均単価(月額)=月額150÷(工数/100)。請求0%なので実想定100で試算→150万 → xlsxの平均単価列で確認
    wb = openpyxl.load_workbook(io.BytesIO(webapp.build_deliveries_xlsx(con)))
    ws = wb["Delivery一覧"]
    assert ws.cell(2, 17).value == 150   # 平均単価(月額・万円/100%)（事業種別L1/L2列の追加で列位置が後ろへ移動）


def test_reschedule_delivery_assignments_slide_and_extend(con):
    """デリバリー期間変更→各アサインの週が連動スライド（#75）。
    開始移動＝全員まるごとスライド／終了のみ移動＝全員の終了だけ延長。"""
    acc = con.execute("INSERT INTO accounts(name) VALUES('社')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="D", stage="受注", status="open")
    dvid = sfa_db.create_delivery(con, deal_id=did, start_week="2026-08-03", end_week="2026-08-31")
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="A",
                                   from_week="2026-08-03", to_week="2026-08-17", fte_pct=100)
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="B",
                                   from_week="2026-08-10", to_week="2026-08-31", fte_pct=50)
    # スライド: 開始+2週（終了も+2週で週数不変）→ 全員 from/to +2週
    sfa_db.reschedule_delivery_assignments(con, dvid, "2026-08-03", "2026-08-31",
                                           "2026-08-17", "2026-09-14")
    byo = {a["owner"]: a for a in sfa_db.list_delivery_assignments(con, dvid)}
    assert (byo["A"]["from_week"], byo["A"]["to_week"]) == ("2026-08-17", "2026-08-31")
    assert (byo["B"]["from_week"], byo["B"]["to_week"]) == ("2026-08-24", "2026-09-14")
    # 週数延長: 開始そのまま・終了+1週 → 全員の終了だけ +1週
    sfa_db.reschedule_delivery_assignments(con, dvid, "2026-08-17", "2026-09-14",
                                           "2026-08-17", "2026-09-21")
    byo = {a["owner"]: a for a in sfa_db.list_delivery_assignments(con, dvid)}
    assert (byo["A"]["from_week"], byo["A"]["to_week"]) == ("2026-08-17", "2026-09-07")
    assert (byo["B"]["from_week"], byo["B"]["to_week"]) == ("2026-08-24", "2026-09-21")
    # 変化なしは0件
    assert sfa_db.reschedule_delivery_assignments(con, dvid, "2026-08-17", "2026-09-21",
                                                  "2026-08-17", "2026-09-21") == 0


def test_add_assignment_defaults_weeks_from_delivery_period(con):
    """週未入力でアサイン追加すると、デリバリーの開始/終了週がデフォルト採用される（webapp経路の要点）。"""
    from cowork import webapp  # noqa: F401 (ルート挙動はweb層。ここではDB既定値の前提を確認)
    acc = con.execute("INSERT INTO accounts(name) VALUES('社2')").lastrowid
    con.commit()
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="D2", stage="受注", status="open")
    dvid = sfa_db.create_delivery(con, deal_id=did, start_week="2026-08-03", end_week="2026-08-31")
    dv = sfa_db.get_delivery(con, dvid)
    # webappの/assignment/addは from/to 未入力時 dv.start_week/end_week を採用する（本体はこの前提）
    assert dv["start_week"] == "2026-08-03" and dv["end_week"] == "2026-08-31"


def test_base_max_periods_effective_and_replace(con):
    """ベース最大稼働率の期間版（#75）: 期間別に実効値を返す・総入れ替え・空行スキップ。"""
    con.execute("INSERT INTO accounts(name) VALUES('社')"); con.commit()
    # 7/6〜8/2=100 / 8/3〜継続=80
    sfa_db.replace_base_max_periods(con, "早瀬", [
        {"from_week": "2026-07-06", "to_week": "2026-08-02", "max_pct": 100},
        {"from_week": "2026-08-03", "to_week": "", "max_pct": 80},
    ])
    ps = sfa_db.list_base_max_periods(con)["早瀬"]
    assert len(ps) == 2 and ps[1]["to_week"] == ""
    assert sfa_db.base_max_at(con, "早瀬", "2026-07-13") == 100
    assert sfa_db.base_max_at(con, "早瀬", "2026-08-10") == 80   # 継続期間
    assert sfa_db.base_max_at(con, "未登録", "2026-08-10") == 100  # 未設定は100
    # 全空行はスキップ（保存されない）
    sfa_db.replace_base_max_periods(con, "中島", [{"from_week": "", "to_week": "", "max_pct": ""}])
    assert not sfa_db.list_base_max_periods(con).get("中島")
    # 総入れ替え（1期間に置換）
    sfa_db.replace_base_max_periods(con, "早瀬", [{"from_week": "", "to_week": "", "max_pct": 50}])
    assert len(sfa_db.list_base_max_periods(con)["早瀬"]) == 1
    assert sfa_db.base_max_at(con, "早瀬", "2027-01-01") == 50   # 開区間=常に適用


def test_delivery_cost_fields_persist_and_convert(con, acc_id):
    """外注費（ユーザー要望2026-08-23）: 報酬額と同じ月額/総額の相互換算＋外注先名の記録。"""
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")  # 4週=1.0ヶ月
    sfa_db.update_delivery(con, dvid, cost_mode="monthly", cost_monthly=50, cost_total=50,
                            cost_vendor="A社")
    dv = sfa_db.get_delivery(con, dvid)
    assert dv["cost_vendor"] == "A社"
    assert sfa_db.delivery_display_costs(dv) == (50, 50)

    # 総額モードで片方だけ入力→月数から補完される
    sfa_db.update_delivery(con, dvid, cost_mode="total", cost_monthly=None, cost_total=200)
    dv2 = sfa_db.get_delivery(con, dvid)
    assert sfa_db.delivery_display_costs(dv2) == (200.0, 200)  # 200/1.0ヶ月=200/月


def test_delivery_profit_is_fee_minus_cost(con, acc_id):
    """想定利益＝報酬額－外注費。外注費未入力なら報酬額そのまま。両方未入力ならNone。"""
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")
    dv0 = sfa_db.get_delivery(con, dvid)
    assert sfa_db.delivery_profit(dv0) == (None, None)

    sfa_db.update_delivery(con, dvid, fee_mode="monthly", fee_monthly=150, fee_total=150)
    dv1 = sfa_db.get_delivery(con, dvid)
    assert sfa_db.delivery_profit(dv1) == (150, 150)  # 外注費未入力=0扱い

    sfa_db.update_delivery(con, dvid, cost_mode="monthly", cost_monthly=60, cost_total=60)
    dv2 = sfa_db.get_delivery(con, dvid)
    assert sfa_db.delivery_profit(dv2) == (90, 90)


def test_delivery_form_renders_cost_fields_and_profit_display(con, acc_id):
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")
    sfa_db.update_delivery(con, dvid, fee_mode="monthly", fee_monthly=150, fee_total=150,
                            cost_mode="monthly", cost_monthly=60, cost_total=60, cost_vendor="B社")
    html = webapp.delivery_form(con, dvid)
    assert 'name="cost_vendor"' in html and 'value="B社"' in html
    assert 'name="cost_mode"' in html
    assert 'id="dvCostMonthly"' in html and 'id="dvCostTotal"' in html
    assert 'id="dvProfitMonthly"' in html and 'id="dvProfitTotal"' in html


def test_delivery_month_range_extends_by_cycle(con, acc_id):
    """delivery_month_range: 開始週〜終了週の月初リスト＋extra_months分の延長（ユーザー要望2026-08-23:
    月別入金計画。検収額と支払いサイクルから入金月を算出するための月範囲展開）。"""
    dv = {"start_week": "2026-09-07", "end_week": "2026-11-30"}
    assert sfa_db.delivery_month_range(dv) == ["2026-09", "2026-10", "2026-11"]
    assert sfa_db.delivery_month_range(dv, extra_months=2) == \
        ["2026-09", "2026-10", "2026-11", "2026-12", "2027-01"]
    assert sfa_db.delivery_month_range({}) == []


def test_delivery_receipt_set_and_delete(con, acc_id):
    """set_delivery_receipt: 保存・上書き・空入力での削除（未入力に戻す）。"""
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 100)
    assert sfa_db.list_delivery_receipts(con, dvid)[0]["amount"] == 100
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 150)  # 上書き
    rows = sfa_db.list_delivery_receipts(con, dvid)
    assert len(rows) == 1 and rows[0]["amount"] == 150
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", "")  # 空入力で削除
    assert sfa_db.list_delivery_receipts(con, dvid) == []
    sfa_db.set_delivery_receipt(con, dvid, "不正な月", 100)  # 不正な月は無視
    assert sfa_db.list_delivery_receipts(con, dvid) == []


def test_delivery_cashflow_shifts_receipts_by_payment_cycle(con, acc_id):
    """delivery_cashflow: 検収額をpayment_cycle_months分ずらして入金額を算出。既定は翌月(1)。"""
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 100)
    sfa_db.set_delivery_receipt(con, dvid, "2026-10", 50)
    cf = sfa_db.delivery_cashflow(con, dvid)
    assert cf["receipts"] == {"2026-09": 100, "2026-10": 50}
    assert cf["payments"] == {"2026-10": 100, "2026-11": 50}  # 既定サイクル=1ヶ月後
    assert "2026-11" in cf["months"]  # 入金月も月一覧に含まれる（テーブル表示のため）

    sfa_db.update_delivery(con, dvid, payment_cycle_months=0)  # 検収月内に入金
    cf2 = sfa_db.delivery_cashflow(con, dvid)
    assert cf2["payments"] == {"2026-09": 100, "2026-10": 50}

    sfa_db.update_delivery(con, dvid, payment_cycle_months=2)  # 検収月+2ヶ月後・年跨ぎ確認は別途
    cf3 = sfa_db.delivery_cashflow(con, dvid)
    assert cf3["payments"] == {"2026-11": 100, "2026-12": 50}


def test_delivery_form_renders_cashflow_table(con, acc_id):
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 100)
    html = webapp.delivery_form(con, dvid)
    assert 'id="dvCashflow"' in html
    assert "2026/09" in html and "2026/10" in html  # _fmt_month表示
    assert 'onchange="dvSetCycle(' in html
    assert f'onchange="dvReceiptSet({dvid},\'2026-09\',this.value)"' in html


def test_delivery_business_type_inherits_from_deal_by_default(con, acc_id):
    """事業種別L1/L2はデフォルトで紐づく商談を継承（ユーザー要望2026-08-23）。"""
    d = sfa_db.upsert_deal(con, account_id=acc_id, deal_name="D", stage="受注",
                           business_type_l1="コスト削減", business_type_l2="コスト診断(無償)")
    dvid = sfa_db.create_delivery(con, deal_id=d)
    dv = sfa_db.get_delivery(con, dvid)
    assert dv["deal_business_type_l1"] == "コスト削減"
    assert sfa_db.delivery_business_type_effective(dv) == ("コスト削減", "コスト診断(無償)")


def test_delivery_business_type_override_takes_precedence(con, acc_id):
    d = sfa_db.upsert_deal(con, account_id=acc_id, deal_name="D", stage="受注",
                           business_type_l1="コスト削減", business_type_l2="コスト診断(無償)")
    dvid = sfa_db.create_delivery(con, deal_id=d)
    sfa_db.update_delivery(con, dvid, business_type_l1_override="コンサルティング")
    dv = sfa_db.get_delivery(con, dvid)
    l1, l2 = sfa_db.delivery_business_type_effective(dv)
    assert l1 == "コンサルティング"


def test_delivery_form_renders_business_type_override_selects(con, acc_id):
    d = sfa_db.upsert_deal(con, account_id=acc_id, deal_name="D", stage="受注",
                           business_type_l1="コスト削減", business_type_l2="コスト診断(無償)")
    dvid = sfa_db.create_delivery(con, deal_id=d)
    html = webapp.delivery_form(con, dvid)
    assert 'name="business_type_l1_override"' in html
    assert 'name="business_type_l2_override"' in html
    # 2026-08-27: 「自動（現在: X）」→「X（自動: 商談を継承）」に表記順を変更（値が先頭に出る）
    assert "コスト削減（自動: 商談を継承）" in html
    assert "コスト診断(無償)（自動: 商談を継承）" in html


def test_build_deliveries_xlsx_includes_business_type_columns(con, acc_id):
    d = sfa_db.upsert_deal(con, account_id=acc_id, deal_name="D", stage="受注",
                           business_type_l1="コスト削減", business_type_l2="コスト診断(無償)")
    dvid = sfa_db.create_delivery(con, deal_id=d)
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(webapp.build_deliveries_xlsx(con)))
    ws = wb["Delivery一覧"]
    hdr = [c.value for c in ws[1]]
    assert "事業種別L1" in hdr and "事業種別L2" in hdr
    row = [c.value for c in ws[2]]
    row_dict = dict(zip(hdr, row))
    assert row_dict["ID"] == dvid
    assert row_dict["事業種別L1"] == "コスト削減"
    assert row_dict["事業種別L2"] == "コスト診断(無償)"


def test_build_deliveries_xlsx_includes_cashflow_sheet(con, acc_id):
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, start_week="2026-09-07", end_week="2026-10-04")
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 100)
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(webapp.build_deliveries_xlsx(con)))
    assert "月別入金計画" in wb.sheetnames
    ws = wb["月別入金計画"]
    assert ws.cell(row=1, column=1).value == "Delivery ID"
    rows = [tuple(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    assert any(r[0] == dvid and r[3] == "2026-09" and r[4] == 100 for r in rows)


def test_payment_schedule_xlsx_receipt_mode_pivots_by_month(con, acc_id):
    """#115: 検収/入金はモードで切替、1案件1行×月別金額の一覧表になること。"""
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, title="A社支援", start_week="2026-09-07",
                                  end_week="2026-11-01", status="進行中")
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 100)
    sfa_db.set_delivery_receipt(con, dvid, "2026-10", 200)
    sfa_db.add_delivery_assignment(con, delivery_id=dvid, owner="早瀬", from_week="2026-09-07",
                                   to_week="2026-09-14", role="コンサルタント", fte_pct=50)
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(webapp.build_delivery_payment_schedule_xlsx(con, mode="receipt")))
    ws = wb.active
    assert ws.title == "検収ベース"
    hdr = [c.value for c in ws[1]]
    assert hdr[:7] == ["#", "クライアント", "案件", "状態", "開始週", "終了週", "アサイン"]
    assert "26/09" in hdr and "26/10" in hdr
    row = [c.value for c in ws[2]]
    row_map = dict(zip(hdr, row))
    assert row_map["#"] == dvid and row_map["案件"] == "A社支援" and row_map["アサイン"] == "早瀬"
    assert row_map["26/09"] == 100 and row_map["26/10"] == 200


def test_payment_schedule_xlsx_payment_mode_shifts_by_cycle(con, acc_id):
    d = _deal(con, acc_id, "受注")
    dvid = sfa_db.create_delivery(con, deal_id=d, title="B社支援", start_week="2026-09-07",
                                  end_week="2026-10-04")
    sfa_db.update_delivery(con, dvid, payment_cycle_months=2)
    sfa_db.set_delivery_receipt(con, dvid, "2026-09", 100)
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(webapp.build_delivery_payment_schedule_xlsx(con, mode="payment")))
    ws = wb.active
    assert ws.title == "入金ベース"
    hdr = [c.value for c in ws[1]]
    row_map = dict(zip(hdr, [c.value for c in ws[2]]))
    assert row_map.get("26/09") is None  # 検収額の月自体は入金ベースには出ない
    assert row_map["26/11"] == 100  # 2ヶ月後に入金


def test_payment_schedule_xlsx_skips_deliveries_without_any_amount(con, acc_id):
    """検収も入金も1件も登録が無い案件は行として出さない（空行で埋めない）。"""
    d = _deal(con, acc_id, "受注")
    sfa_db.create_delivery(con, deal_id=d, title="登録なし案件")
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(webapp.build_delivery_payment_schedule_xlsx(con, mode="receipt")))
    ws = wb.active
    assert ws.max_row == 1  # ヘッダのみ


# ── Delivery複製（ユーザー要望2026-08-27） ────────────────────────────────

def test_duplicate_delivery_copies_plan_fields_but_not_execution_data(con, acc_id):
    """deal_duplicateと同じ思想: 体制(目標役割)・報酬/外注費設定は引き継ぐが、
    実行済みのアサイン実績・検収実額・確度の手動固定は引き継がず、真っ白から始める。"""
    d = _deal(con, acc_id, "受注")
    src_id = sfa_db.create_delivery(con, deal_id=d, title="A社支援", start_week="2026-09-07",
                                    end_week="2026-10-04", status="完了",
                                    overview="概要テキスト", confidence_override="確定")
    sfa_db.update_delivery(con, src_id, fee_mode="monthly", fee_monthly=100, cost_mode="monthly",
                           cost_monthly=20, cost_vendor="外注先X", payment_cycle_months=2,
                           business_type_l1_override="コスト削減", business_type_l2_override="診断")
    sfa_db.add_delivery_role(con, delivery_id=src_id, role="リード", fte_billing=15, fte_pct=5)
    sfa_db.add_delivery_role(con, delivery_id=src_id, role="コンサルタント", fte_billing=50, fte_pct=30)
    sfa_db.add_delivery_assignment(con, delivery_id=src_id, owner="高橋", from_week="2026-09-07",
                                   to_week="2026-09-14", role="コンサルタント", fte_pct=50)
    sfa_db.set_delivery_receipt(con, src_id, "2026-09", 100)

    new_id = sfa_db.duplicate_delivery(con, src_id)
    assert new_id and new_id != src_id
    new = sfa_db.get_delivery(con, new_id)
    assert new["title"] == "A社支援（コピー）"
    assert new["deal_id"] == d
    assert new["status"] == "進行中"  # 真っ白から始める
    assert new["confidence_override"] is None  # 手動固定は引き継がない→自動導出に戻る
    assert new["overview"] == "概要テキスト"
    assert new["fee_mode"] == "monthly" and new["fee_monthly"] == 100
    assert new["cost_vendor"] == "外注先X"
    assert new["payment_cycle_months"] == 2
    assert new["business_type_l1_override"] == "コスト削減"

    new_roles = sfa_db.list_delivery_roles(con, new_id)
    assert {(r["role"], r["fte_billing"], r["fte_pct"]) for r in new_roles} == {
        ("リード", 15.0, 5.0), ("コンサルタント", 50.0, 30.0)}

    # アサイン実績・検収実額はコピーしない
    assert sfa_db.list_delivery_assignments(con, new_id) == []
    assert sfa_db.list_delivery_receipts(con, new_id) == []
    # 元Deliveryは変更されない
    assert sfa_db.list_delivery_assignments(con, src_id) != []


def test_duplicate_delivery_missing_source_returns_none(con):
    assert sfa_db.duplicate_delivery(con, 999999) is None


def test_delivery_confidence_and_business_type_auto_option_label_shows_value_first(con, acc_id):
    """自動判定の選択肢は「自動（現在: X）」ではなく「X（自動判定）」の順で表示する
    （ユーザー要望2026-08-27: カッコの内外が逆で分かりにくいとの指摘）。"""
    did = sfa_db.upsert_deal(con, account_id=acc_id, deal_name="案件X", stage="受注", status="open")
    con.commit()
    dv = sfa_db.get_delivery(con, sfa_db.create_delivery(con, deal_id=did, title="納品X"))

    html = webapp.delivery_form(con, dv["id"])
    assert "自動（現在" not in html
    assert "確定（自動判定）" in html  # stage=受注→自動判定は「確定」が先頭に出る

    l1_opts = webapp._delivery_biz_l1_opts(con, dv)
    l2_opts = webapp._delivery_biz_l2_opts(con, dv)
    assert "自動（商談を継承" not in l1_opts and "自動（商談を継承" not in l2_opts
    assert "（自動: 商談を継承）" in l1_opts and "（自動: 商談を継承）" in l2_opts
