"""開発要件一覧(#165, 2026-09-04)の回帰テスト。

ユーザー確定仕様:
- dev_projects(営業側の開発専用機能)とは別の新規テーブル。
- 商談が「提案」以降ステージに到達、またはDeliveryが作成された時点で1行を自動起票
  （詳細は空欄=dev_involved未判定）。同一商談が両トリガーを踏むと2行できるケースを許容する。
- 「開発有無=無」の行は識別情報(紐づけ先/アカウント/案件名/ステージ/営業主担当)以外を
  "-"表示にする（一覧・xlsx出力の両方）。
- xlsx出力（/dev-requirements/export.xlsx）。
- 初回バックフィルはscripts/backfill_dev_requirements.pyで実施し、
  「プロジェクト概要・ゴール」はHaiku（現状メモ＋直近活動履歴が入力）で自動生成する。
"""
from __future__ import annotations

import base64
import shutil
import sys
import tempfile
import urllib.error
import urllib.request
from http.server import ThreadingHTTPServer
from pathlib import Path

import pytest

from cowork import sfa_db, webapp

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import scripts.backfill_dev_requirements as backfill_mod

BASIC_USER = "test_user"
BASIC_PASS = "test_pass_1234"


@pytest.fixture
def con():
    d = tempfile.mkdtemp(prefix="sfa_dev_req_")
    path = str(Path(d) / "t.db")
    sfa_db.init_db(path)
    conn = sfa_db.connect(path)
    yield conn
    conn.close()
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def server(monkeypatch, tmp_path):
    db_path = str(tmp_path / "srv.db")
    sfa_db.init_db(db_path)
    monkeypatch.setattr(webapp, "SFA_BASIC_USER", BASIC_USER)
    monkeypatch.setattr(webapp, "SFA_BASIC_PASS", BASIC_PASS)
    handler_cls = webapp._make_handler(db_path, None)
    srv = ThreadingHTTPServer(("127.0.0.1", 0), handler_cls)
    port = srv.server_address[1]
    import threading
    t = threading.Thread(target=srv.serve_forever, daemon=True)
    t.start()
    try:
        yield f"http://127.0.0.1:{port}"
    finally:
        srv.shutdown()
        srv.server_close()
        t.join(timeout=5)


def _auth_header():
    token = base64.b64encode(f"{BASIC_USER}:{BASIC_PASS}".encode()).decode()
    return {"Authorization": f"Basic {token}"}


def _get_bytes(url):
    req = urllib.request.Request(url, headers=_auth_header(), method="GET")
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        return resp.getcode(), resp.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()


def _post_form(url, data: dict):
    import urllib.parse as up
    body = up.urlencode(data, doseq=True).encode()
    req = urllib.request.Request(url, data=body, headers=_auth_header(), method="POST")
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        return resp.getcode(), resp.geturl(), resp.read()
    except urllib.error.HTTPError as e:
        return e.code, url, e.read()


def _post_multipart_file(url, field_name, filename, content: bytes):
    boundary = "----testboundary123"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"\r\n'
        "Content-Type: application/octet-stream\r\n\r\n"
    ).encode() + content + f"\r\n--{boundary}--\r\n".encode()
    headers = dict(_auth_header())
    headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        return resp.getcode(), resp.geturl(), resp.read()
    except urllib.error.HTTPError as e:
        return e.code, url, e.read()


# ── sfa_db層: トリガー・CRUD ──

def test_ensure_dev_requirement_for_deal_only_fires_on_trigger_stages(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="要件詰め", status="open")
    assert sfa_db.ensure_dev_requirement_for_deal(con, did, "要件詰め") is None
    assert sfa_db.get_dev_requirement_by_link(con, "deal", did) is None

    assert sfa_db.ensure_dev_requirement_for_deal(con, did, "提案") is not None
    row = sfa_db.get_dev_requirement_by_link(con, "deal", did)
    assert row is not None
    assert row["dev_involved"] is None  # 未判定


def test_ensure_dev_requirement_for_deal_is_idempotent(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid1 = sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did, "クロージング")
    assert rid1 is not None
    assert rid2 is None
    assert len(sfa_db.list_dev_requirements(con)) == 1


def test_create_delivery_auto_creates_dev_requirement(con):
    """Delivery作成時は必ず開発要件行が自動起票される（create_delivery内に集約、#165）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="DeliveryX")
    row = sfa_db.get_dev_requirement_by_link(con, "delivery", dv_id)
    assert row is not None
    assert row["dev_involved"] is None


def test_deal_and_delivery_triggers_both_fire_creates_two_db_rows_but_list_shows_one(con):
    """同一商談が「提案到達」と「Delivery作成」の両方を踏むと、DBには2行できる
    （トリガー自体の仕様は変更なし・ユーザー確定）。ただし一覧・xlsxの表示は分かりづらいため
    Delivery起点行に集約し、商談起点行は非表示にする（2026-09-06追補・ユーザー確定）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    sfa_db.create_delivery(con, deal_id=did, title="DeliveryX")

    raw_rows = con.execute("SELECT * FROM dev_requirements").fetchall()
    assert len(raw_rows) == 2  # DB上は2行のまま

    rows = sfa_db.list_dev_requirements(con)
    assert len(rows) == 1
    assert rows[0]["link_type"] == "delivery"


def test_deal_row_reappears_in_list_if_delivery_is_deleted(con):
    """集約はあくまで表示上の話。Deliveryが削除されれば、商談起点行は再び一覧に出る
    （DB上の商談起点行自体は削除されないため）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="DeliveryX")
    assert len(sfa_db.list_dev_requirements(con)) == 1

    sfa_db.delete_delivery(con, dv_id)
    rows = sfa_db.list_dev_requirements(con)
    assert len(rows) == 1
    assert rows[0]["link_type"] == "deal"


def test_deal_row_shown_when_delivery_belongs_to_different_deal(con):
    """別の商談のDeliveryが存在するだけでは、この商談の商談起点行は隠されない。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件Y", stage="受注", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    sfa_db.create_delivery(con, deal_id=did2, title="DeliveryY")

    rows = sfa_db.list_dev_requirements(con)
    assert len(rows) == 2
    assert {r["link_type"] for r in rows} == {"deal", "delivery"}


def test_set_dev_requirement_field_does_not_clobber_other_fields(con):
    """set_dev_requirement_fieldは1列だけを更新し、他の既存値を消さない
    （upsert_task/upsert_deal_issueの全列上書きfootgunを避けるための専用ヘルパー）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid = sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    sfa_db.set_dev_requirement_field(con, rid, "overview", "概要A")
    sfa_db.set_dev_requirement_field(con, rid, "budget", "500万円")
    row = sfa_db.get_dev_requirement(con, rid)
    assert row["overview"] == "概要A"
    assert row["budget"] == "500万円"


def test_list_dev_requirements_enriches_with_deal_and_delivery_context(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案",
                             status="open", owner="早瀬")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    rows = sfa_db.list_dev_requirements(con)
    assert len(rows) == 1
    assert rows[0]["account_name"] == "A社"
    assert rows[0]["project_name"] == "案件X"
    assert rows[0]["stage"] == "提案"
    assert rows[0]["owner"] == "早瀬"


def test_list_dev_requirements_delivery_row_pulls_dates_from_delivery(con):
    """#165追補(2026-09-04ユーザー要望): Delivery起点の行はプロジェクト開始日/終了日を
    deliveries.start_week/end_weekから自動反映する。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="DeliveryX",
                                   start_week="2026-09-07", end_week="2026-10-05")
    rows = sfa_db.list_dev_requirements(con)
    row = next(r for r in rows if r["link_type"] == "delivery")
    assert row["start_date"] == "2026-09-07"
    assert row["end_date"] == "2026-10-05"


def test_list_dev_requirements_delivery_row_manual_override_wins(con):
    """開発要件行にstart_date/end_dateを人間が明示入力していれば、そちらを優先する。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="DeliveryX",
                                   start_week="2026-09-07", end_week="2026-10-05")
    rid = sfa_db.get_dev_requirement_by_link(con, "delivery", dv_id)["id"]
    sfa_db.set_dev_requirement_field(con, rid, "start_date", "2026-08-01")
    rows = sfa_db.list_dev_requirements(con)
    row = next(r for r in rows if r["link_type"] == "delivery")
    assert row["start_date"] == "2026-08-01"
    assert row["end_date"] == "2026-10-05"  # 上書きしていない方はDelivery側の値のまま


def test_list_dev_requirements_pulls_demo_link_from_dev_project(con):
    """#165追補(2026-09-04ユーザー要望): デモリンクは、その商談に紐づく開発案件の
    「制作したツールのリンク」があれば自動反映する（商談起点・Delivery起点の両方）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    sfa_db.upsert_dev_project(con, deal_id=did, theme="デモ", status="開発中", stage="プロト",
                              tool_url="https://example.com/demo")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "受注")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="DeliveryX")

    rows = sfa_db.list_dev_requirements(con)
    assert {r.get("demo_link") for r in rows} == {"https://example.com/demo"}


def test_list_dev_requirements_demo_link_manual_override_wins(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    sfa_db.upsert_dev_project(con, deal_id=did, theme="デモ", status="開発中", stage="プロト",
                              tool_url="https://example.com/demo")
    rid = sfa_db.ensure_dev_requirement_for_deal(con, did, "受注")
    sfa_db.set_dev_requirement_field(con, rid, "demo_link", "https://example.com/manual")

    row = sfa_db.get_dev_requirement(con, rid)
    rows = sfa_db.list_dev_requirements(con)
    assert next(r for r in rows if r["id"] == rid)["demo_link"] == "https://example.com/manual"


def test_list_dev_requirements_demo_link_blank_when_no_dev_project(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "受注")
    rows = sfa_db.list_dev_requirements(con)
    assert rows[0].get("demo_link") is None


def test_list_dev_requirements_excludes_rows_with_deleted_link_target(con):
    """紐づけ先の商談が削除されると、開発要件一覧からも除外される（参照切れガード）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    con.execute("DELETE FROM deals WHERE id=?", (did,))
    con.commit()
    assert sfa_db.list_dev_requirements(con) == []


def test_dev_involved_options_and_masters_registered(con):
    assert sfa_db.get_master_list(con, "dev_req_confidentiality") == sfa_db.DEV_REQUIREMENT_CONFIDENTIALITY_LEVELS
    assert sfa_db.get_master_list(con, "dev_req_contract_types") == sfa_db.DEV_REQUIREMENT_CONTRACT_TYPES
    assert sfa_db.get_master_list(con, "dev_req_statuses") == sfa_db.DEV_REQUIREMENT_STATUSES


# ── webapp層: xlsx出力 ──

def test_build_dev_requirements_xlsx_masks_fields_when_no_dev(con):
    import io
    import openpyxl

    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid1 = sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    sfa_db.set_dev_requirement_field(con, rid1, "overview", "開発ありの概要")

    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件Y", stage="提案", status="open")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "提案")
    sfa_db.set_dev_requirement_field(con, rid2, "dev_involved", "無")
    sfa_db.set_dev_requirement_field(con, rid2, "overview", "これは表示されないはず")

    xls = webapp.build_dev_requirements_xlsx(con)
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    overview_idx = header.index("プロジェクト概要・ゴール")
    account_idx = header.index("アカウント")
    project_idx = header.index("案件名")

    by_project = {r[project_idx]: r for r in rows[1:]}
    assert by_project["案件X"][overview_idx] == "開発ありの概要"
    assert by_project["案件Y"][overview_idx] == "-"
    # 識別情報は「開発有無=無」でもマスクされない
    assert by_project["案件Y"][account_idx] == "A社"


def test_build_dev_requirements_xlsx_sets_column_widths(con):
    import io
    import openpyxl

    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")

    xls = webapp.build_dev_requirements_xlsx(con)
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    ws = wb.active

    for c, (key, _label) in enumerate(webapp.DEV_REQ_XLSX_COLUMNS, 1):
        expected = webapp.DEV_REQ_XLSX_COLUMN_WIDTHS.get(key)
        if expected is None:
            continue
        letter = openpyxl.utils.get_column_letter(c)
        assert ws.column_dimensions[letter].width == expected, key


def test_dev_requirements_export_route(server):
    code, resp = _get_bytes(server + "/dev-requirements/export.xlsx")
    assert code == 200


# ── scripts/backfill_dev_requirements.py ──

def test_backfill_find_candidates_filters_by_stage(con):
    acc = sfa_db.upsert_account(con, name="A社")
    sfa_db.upsert_deal(con, account_id=acc, deal_name="対象外", stage="要件詰め", status="open")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="対象", stage="提案", status="open")
    candidates = backfill_mod.find_candidates(con)
    assert len(candidates) == 1
    assert candidates[0]["link_id"] == did
    assert candidates[0]["link_type"] == "deal"


def test_backfill_skips_already_created_rows(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="対象", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    assert backfill_mod.find_candidates(con) == []


def test_backfill_generate_overview_uses_note_and_activities(con, monkeypatch):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="請求書自動化", stage="提案",
                             status="open", note="請求書処理を自動化したい")
    sfa_db.add_activity(con, deal_id=did, type="面談", occurred_on="2026-09-01",
                        body="受領から支払いまでを自動化したいとの要望")

    captured = {}

    def _fake_haiku(prompt, **kw):
        captured["prompt"] = prompt
        return "生成された概要"

    monkeypatch.setattr(webapp, "_call_claude_haiku", _fake_haiku)
    overview = backfill_mod.generate_overview(con, did)
    assert overview == "生成された概要"
    assert "請求書処理を自動化したい" in captured["prompt"]
    assert "受領から支払いまでを自動化したい" in captured["prompt"]


def test_backfill_generate_overview_empty_when_no_deal(con):
    assert backfill_mod.generate_overview(con, None) == ""


def test_backfill_apply_creates_rows_and_sets_overview(con, monkeypatch):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open",
                             note="要件は請求書自動化の相談")
    monkeypatch.setattr(webapp, "_call_claude_haiku", lambda prompt, **kw: "自動生成された概要")

    candidates = backfill_mod.find_candidates(con)
    n = backfill_mod.apply_backfill(con, candidates)
    assert n == 1
    row = sfa_db.get_dev_requirement_by_link(con, "deal", did)
    assert row["overview"] == "自動生成された概要"


def test_backfill_apply_skips_haiku_when_no_material(con, monkeypatch):
    """現状メモ・活動履歴のどちらも無い場合、Haikuを呼ばずoverviewは空欄のまま
    （2026-09-04ユーザー報告: 材料が無いとHaikuが謝罪文を書いてしまっていた不具合の修正）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    calls = []
    monkeypatch.setattr(webapp, "_call_claude_haiku",
                        lambda prompt, **kw: calls.append(prompt) or "呼ばれないはず")

    candidates = backfill_mod.find_candidates(con)
    backfill_mod.apply_backfill(con, candidates)
    row = sfa_db.get_dev_requirement_by_link(con, "deal", did)
    assert row["overview"] is None
    assert calls == []


def test_backfill_apply_survives_haiku_failure(con, monkeypatch):
    """Haiku生成が失敗しても、行の作成自体は失敗しない（overview空欄で継続）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open",
                       note="請求書自動化を検討中")

    def _boom(prompt, **kw):
        raise RuntimeError("API error")
    monkeypatch.setattr(webapp, "_call_claude_haiku", _boom)

    candidates = backfill_mod.find_candidates(con)
    n = backfill_mod.apply_backfill(con, candidates)
    assert n == 1
    rows = sfa_db.list_dev_requirements(con)
    assert rows[0]["overview"] is None


# ── scripts/backfill_dev_requirements.py --fix-overviews ──

def test_find_apology_overview_rows_detects_marker_text(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="正常", stage="提案", status="open")
    rid1 = sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    sfa_db.set_dev_requirement_field(con, rid1, "overview", "正常な概要文")

    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="謝罪混入", stage="提案", status="open")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "提案")
    sfa_db.set_dev_requirement_field(con, rid2, "overview", "申し訳ございませんが、情報が不足しております。")

    rows = backfill_mod.find_apology_overview_rows(con)
    assert [r["id"] for r in rows] == [rid2]


def test_fix_overviews_regenerates_with_material_and_clears_without(con, monkeypatch):
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="材料あり", stage="提案", status="open",
                              note="請求書自動化の相談")
    rid1 = sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    sfa_db.set_dev_requirement_field(con, rid1, "overview", "申し訳ございませんが情報不足です")

    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="材料なし", stage="提案", status="open")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "提案")
    sfa_db.set_dev_requirement_field(con, rid2, "overview", "申し訳ございませんが情報不足です")

    monkeypatch.setattr(webapp, "_call_claude_haiku", lambda prompt, **kw: "再生成された概要")

    rows = backfill_mod.find_apology_overview_rows(con)
    assert len(rows) == 2
    n = backfill_mod.fix_overviews(con, rows)
    assert n == 2

    assert sfa_db.get_dev_requirement(con, rid1)["overview"] == "再生成された概要"
    assert sfa_db.get_dev_requirement(con, rid2)["overview"] is None  # 材料なし=Haiku呼ばず空欄に戻る


def test_resolve_deal_id_for_delivery_row(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    dv_id = sfa_db.create_delivery(con, deal_id=did, title="DeliveryX")
    row = sfa_db.get_dev_requirement_by_link(con, "delivery", dv_id)
    assert backfill_mod._resolve_deal_id(con, row) == did


# ── #165画面フェーズ1: 一覧ページ ──

def test_dev_requirements_page_lists_rows_and_masks_no_dev_row(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件Y", stage="提案", status="open")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "提案")
    sfa_db.set_dev_requirement_field(con, rid2, "dev_involved", "無")
    sfa_db.set_dev_requirement_field(con, rid2, "overview", "これは表示されないはず")

    html = webapp.dev_requirements_page(con)
    assert "案件X" in html
    assert "案件Y" in html
    assert "これは表示されないはず" not in html


def test_dev_requirements_page_filters_by_dev_involved(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件Y", stage="提案", status="open")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "提案")
    sfa_db.set_dev_requirement_field(con, rid2, "dev_involved", "無")

    html_no = webapp.dev_requirements_page(con, dev_involved="無")
    assert "案件Y" in html_no
    assert "案件X" not in html_no

    html_mikan = webapp.dev_requirements_page(con, dev_involved="未判定")
    assert "案件X" in html_mikan
    assert "案件Y" not in html_mikan


def test_dev_requirements_page_shows_empty_message_when_no_rows(con):
    html = webapp.dev_requirements_page(con)
    assert "該当する開発要件がありません" in html


def test_dev_requirements_page_demo_link_and_account_link(con):
    """デモリンクは🔗アイコン、アカウント名は紐づく商談/Deliveryへのリンクになる（#165画面設計確定仕様）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="受注", status="open")
    sfa_db.upsert_dev_project(con, deal_id=did, theme="デモ", status="開発中", stage="プロト",
                              tool_url="https://example.com/demo")
    rid = sfa_db.ensure_dev_requirement_for_deal(con, did, "受注")

    html = webapp.dev_requirements_page(con)
    assert f'href="/deal/{did}"' in html
    assert 'href="https://example.com/demo"' in html
    assert "🔗" in html


def test_dev_requirements_route_via_http(server):
    code, body = _get_bytes(f"{server}/dev-requirements")
    assert code == 200
    assert "開発要件一覧".encode() in body


# ── #165フェーズ2: インライン編集 ──

def test_dev_requirements_page_renders_editable_controls(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid = sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    html = webapp.dev_requirements_page(con)
    assert f'data-dr-id="{rid}"' in html
    assert 'data-field="dev_involved"' in html
    assert 'data-field="release_target"' in html
    assert f'id="dr-detail-{rid}"' in html
    assert "drToggle(" in html


def test_dev_requirement_field_route_via_db(con, monkeypatch, tmp_path):
    import base64
    import threading
    from http.server import ThreadingHTTPServer

    db_path = str(tmp_path / "srv3.db")
    sfa_db.init_db(db_path)
    con2 = sfa_db.connect(db_path)
    acc = sfa_db.upsert_account(con2, name="A社")
    did = sfa_db.upsert_deal(con2, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid = sfa_db.ensure_dev_requirement_for_deal(con2, did, "提案")
    con2.close()

    user, pw = "u", "p"
    monkeypatch.setattr(webapp, "SFA_BASIC_USER", user)
    monkeypatch.setattr(webapp, "SFA_BASIC_PASS", pw)
    handler_cls = webapp._make_handler(db_path, None)
    srv = ThreadingHTTPServer(("127.0.0.1", 0), handler_cls)
    port = srv.server_address[1]
    t = threading.Thread(target=srv.serve_forever, daemon=True)
    t.start()
    base = f"http://127.0.0.1:{port}"
    try:
        import urllib.request as ur
        tok = base64.b64encode(f"{user}:{pw}".encode()).decode()

        def post(url, data):
            import urllib.parse as up
            body = up.urlencode(data).encode()
            req = ur.Request(url, data=body, headers={"Authorization": f"Basic {tok}"}, method="POST")
            try:
                resp = ur.urlopen(req, timeout=10)
                return resp.getcode()
            except ur.HTTPError as e:
                return e.code

        code = post(f"{base}/dev-requirement/{rid}/field", {"field": "release_target", "value": "2027年1月"})
        assert code == 204
        code2 = post(f"{base}/dev-requirement/{rid}/field", {"field": "dev_status", "value": "不正なステータス"})
        assert code2 == 400
        code3 = post(f"{base}/dev-requirement/{rid}/field", {"field": "link_type", "value": "delivery"})
        assert code3 == 400  # link_type/link_idはUIから変更不可
    finally:
        srv.shutdown()
        srv.server_close()
        t.join(timeout=5)

    con3 = sfa_db.connect(db_path)
    row = sfa_db.get_dev_requirement(con3, rid)
    assert row["release_target"] == "2027年1月"
    assert row["link_type"] == "deal"  # 変更されていない


# ── #165フェーズ3: テンプレUpload → 差分確認 → 反映 ──

def test_parse_upload_detects_update_via_id_match(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid = sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")
    sfa_db.set_dev_requirement_field(con, rid, "budget", "500万円")

    xls = webapp.build_dev_requirements_xlsx(con)
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    ws = wb.active
    header = [c.value for c in ws[1]]
    budget_col = header.index("予算感") + 1
    ws.cell(row=2, column=budget_col, value="800万円")
    buf = io.BytesIO()
    wb.save(buf)

    diffs, err = webapp._dev_requirements_parse_upload(con, buf.getvalue())
    assert err == ""
    assert len(diffs) == 1
    assert diffs[0]["status"] == "update"
    assert diffs[0]["dev_requirement_id"] == rid
    change = next(c for c in diffs[0]["changes"] if c["field"] == "budget")
    assert change["old"] == "500万円" and change["new"] == "800万円"


def test_parse_upload_no_change_returns_empty_diff(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")

    xls = webapp.build_dev_requirements_xlsx(con)
    diffs, err = webapp._dev_requirements_parse_upload(con, xls)
    assert err == ""
    assert diffs == []


def test_parse_upload_missing_id_columns_returns_error(con):
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["適当な列"])
    buf = io.BytesIO()
    wb.save(buf)
    diffs, err = webapp._dev_requirements_parse_upload(con, buf.getvalue())
    assert diffs == []
    assert "SFA商談#" in err


def test_parse_upload_invalid_file_returns_error(con):
    diffs, err = webapp._dev_requirements_parse_upload(con, b"not an xlsx file")
    assert diffs == []
    assert err


def test_parse_upload_unresolvable_id_is_error_row(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")

    xls = webapp.build_dev_requirements_xlsx(con)
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    ws = wb.active
    header = [c.value for c in ws[1]]
    deal_no_col = header.index("SFA商談#") + 1
    ws.cell(row=2, column=deal_no_col, value=999999)  # 存在しない商談ID
    buf = io.BytesIO()
    wb.save(buf)

    diffs, err = webapp._dev_requirements_parse_upload(con, buf.getvalue())
    assert err == ""
    assert len(diffs) == 1
    assert diffs[0]["status"] == "error"


def test_parse_upload_new_row_for_existing_deal_without_dev_requirement(con):
    """dev_requirements行がまだ無い実在の商談を指すSFA商談#が来たら「新規」として扱う
    （新規の商談自体を作る機能は無い、既存商談への追記のみ、design doc確定）。"""
    acc = sfa_db.upsert_account(con, name="A社")
    did = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid = sfa_db.ensure_dev_requirement_for_deal(con, did, "提案")

    xls = webapp.build_dev_requirements_xlsx(con)
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    ws = wb.active
    header = [c.value for c in ws[1]]
    budget_col = header.index("予算感") + 1
    ws.cell(row=2, column=budget_col, value="900万円")
    buf = io.BytesIO()
    wb.save(buf)

    sfa_db.delete_dev_requirement(con, rid)  # 行を削除し「新規」ケースを再現
    diffs, err = webapp._dev_requirements_parse_upload(con, buf.getvalue())
    assert err == ""
    assert len(diffs) == 1
    assert diffs[0]["status"] == "new"
    assert diffs[0]["link_type"] == "deal" and diffs[0]["link_id"] == did


def test_upload_review_commit_flow_via_http(server, tmp_path):
    """アップロード→レビュー画面→commitの一連の流れをHTTP経由で確認する。"""
    db_path = str(tmp_path / "srv.db")  # serverフィクスチャと同じDBファイル
    con2 = sfa_db.connect(db_path)
    acc = sfa_db.upsert_account(con2, name="A社")
    did = sfa_db.upsert_deal(con2, account_id=acc, deal_name="案件X", stage="提案", status="open")
    rid = sfa_db.ensure_dev_requirement_for_deal(con2, did, "提案")
    xls = webapp.build_dev_requirements_xlsx(con2)
    con2.close()

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xls))
    ws = wb.active
    header = [c.value for c in ws[1]]
    budget_col = header.index("予算感") + 1
    ws.cell(row=2, column=budget_col, value="777万円")
    buf = io.BytesIO()
    wb.save(buf)

    code, url, body = _post_multipart_file(f"{server}/dev-requirements/upload", "xlsx_file",
                                           "dev_requirements.xlsx", buf.getvalue())
    assert code == 200
    assert "/dev-requirements/review/" in url
    assert "予算感".encode() in body
    assert "777万円".encode() in body
    upload_id = int(url.rstrip("/").rsplit("/", 1)[-1])

    code2, _, body2 = _post_form(f"{server}/dev-requirements/commit",
                                 {"upload_id": upload_id, "apply": ["0"]})
    assert code2 == 200

    con3 = sfa_db.connect(db_path)
    assert sfa_db.get_dev_requirement(con3, rid)["budget"] == "777万円"
    assert sfa_db.get_dev_requirements_upload(con3, upload_id) is None


def test_commit_applies_selected_changes_and_skips_unchecked(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件X", stage="提案", status="open")
    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="案件Y", stage="提案", status="open")
    rid1 = sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "提案")

    diffs = [
        {"status": "update", "label": "A社/案件X", "dev_requirement_id": rid1,
         "changes": [{"field": "budget", "label": "予算感", "old": None, "new": "300万円"}]},
        {"status": "update", "label": "A社/案件Y", "dev_requirement_id": rid2,
         "changes": [{"field": "budget", "label": "予算感", "old": None, "new": "400万円"}]},
    ]
    upload_id = sfa_db.create_dev_requirements_upload(con, diffs)

    # apply=0 のみ選択（案件Yは反映しない）を模擬してcommit相当のロジックを直接検証
    stored = sfa_db.get_dev_requirements_upload(con, upload_id)
    assert stored == diffs
    for i in [0]:
        d = stored[i]
        for c in d["changes"]:
            sfa_db.set_dev_requirement_field(con, d["dev_requirement_id"], c["field"], c["new"])
    sfa_db.delete_dev_requirements_upload(con, upload_id)

    assert sfa_db.get_dev_requirement(con, rid1)["budget"] == "300万円"
    assert sfa_db.get_dev_requirement(con, rid2)["budget"] is None
    assert sfa_db.get_dev_requirements_upload(con, upload_id) is None


def test_review_page_missing_upload_shows_message(con):
    html = webapp.dev_requirements_review_page(con, 999999)
    assert "見つかりません" in html


# ── フィルタ機能（2026-09-06追加要望） ──

def test_dev_requirements_page_filters_by_owner_stage_dev_status_and_keyword(con):
    acc = sfa_db.upsert_account(con, name="A社")
    did1 = sfa_db.upsert_deal(con, account_id=acc, deal_name="請求書自動化案件", stage="提案",
                              status="open", owner="早瀬")
    did2 = sfa_db.upsert_deal(con, account_id=acc, deal_name="在庫管理DX", stage="クロージング",
                              status="open", owner="山端")
    rid1 = sfa_db.ensure_dev_requirement_for_deal(con, did1, "提案")
    rid2 = sfa_db.ensure_dev_requirement_for_deal(con, did2, "クロージング")
    sfa_db.set_dev_requirement_field(con, rid1, "dev_status", "検討中")
    sfa_db.set_dev_requirement_field(con, rid2, "dev_status", "未着手")

    html_owner = webapp.dev_requirements_page(con, owner="早瀬")
    assert "請求書自動化案件" in html_owner and "在庫管理DX" not in html_owner

    html_stage = webapp.dev_requirements_page(con, stage="クロージング")
    assert "在庫管理DX" in html_stage and "請求書自動化案件" not in html_stage

    html_status = webapp.dev_requirements_page(con, dev_status="検討中")
    assert "請求書自動化案件" in html_status and "在庫管理DX" not in html_status

    html_q = webapp.dev_requirements_page(con, q="在庫")
    assert "在庫管理DX" in html_q and "請求書自動化案件" not in html_q


def test_dev_requirements_page_renders_filter_form(con):
    html = webapp.dev_requirements_page(con)
    assert 'name="owner"' in html
    assert 'name="stage"' in html
    assert 'name="dev_status"' in html
    assert 'name="q"' in html


def test_dev_requirements_route_via_http_with_filters(server):
    code, body = _get_bytes(f"{server}/dev-requirements?owner=%E6%97%A9%E7%80%A8&stage=%E6%8F%90%E6%A1%88")
    assert code == 200
