"""xlsxのテーマKPIシートからコスト削減モデルのフィールドを商談レコードに流し込む。

対象: 分類(category)=Salesの行のみ。
キー: deals.theme_id ← xlsx.ID
更新列: approach_value, approach_rate, reduction_rate, fee_rate,
        diagnosis_cost, cost_stage, importance

使い方:
    python3 scripts/fill_cost_fields.py            # 実際に書き込む
    python3 scripts/fill_cost_fields.py --dry-run  # 書き込まずに確認だけ
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

# .env 読み込み（restore_meeting_dates.py と同じパターン）
_env_path = ROOT / ".env"
if _env_path.exists():
    for _line in _env_path.read_text().splitlines():
        if "=" in _line and not _line.startswith("#"):
            _k, _, _v = _line.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip())

sys.path.insert(0, str(ROOT))

from cowork import sfa_db
from cowork.mapping import _f, _s, normalize_header

XLSX_PATH = Path("/mnt/c/Users/ninse/OneDrive/個人用/ツール作成/秘書/sales_themes_input.xlsx")
SHEET_NAME = "テーマKPI"
DB_PATH = os.environ.get("COWORK_SFA_DB", sfa_db.DEFAULT_DB_PATH)

# xlsx列名（normalize_header後）→ dealsカラム名
FIELD_MAP = {
    "アプローチ額(億円)": "approach_value",
    "アプローチ率(%)":    "approach_rate",
    "コスト削減率(%)":    "reduction_rate",
    "成果報酬率(%)":      "fee_rate",
    "診断原価(万円)":     "diagnosis_cost",
    "コスト削減ステージ": "cost_stage",
    "重要度":             "importance",
}

# 数値型カラムのセット
NUMERIC_COLS = {"approach_value", "approach_rate", "reduction_rate", "fee_rate", "diagnosis_cost"}


def read_xlsx() -> list[dict]:
    """xlsxのテーマKPIシートを読み込み、Sales行のみを返す。"""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl がインストールされていません。pip install openpyxl を実行してください。")

    if not XLSX_PATH.exists():
        sys.exit(f"xlsx が見つかりません: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"シート '{SHEET_NAME}' が見つかりません。利用可能: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    headers = [normalize_header(c.value) for c in ws[1]]

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue
        r = dict(zip(headers, raw))
        if _s(r.get("分類")) != "Sales":
            continue
        rows.append(r)
    return rows


def build_updates(rows: list[dict]) -> list[tuple[int, dict]]:
    """各行からtheme_idと更新フィールドdict（NULLスキップ済み）を生成する。"""
    result = []
    for r in rows:
        try:
            theme_id = int(r.get("ID") or 0)
        except (ValueError, TypeError):
            continue
        if theme_id <= 0:
            continue

        updates: dict[str, object] = {}
        for xlsx_col, db_col in FIELD_MAP.items():
            raw = r.get(xlsx_col)
            if db_col in NUMERIC_COLS:
                val = _f(raw)
            else:
                val = _s(raw)
            if val is not None:
                updates[db_col] = val

        if updates:
            result.append((theme_id, updates))
    return result


def apply_updates(updates: list[tuple[int, dict]], dry_run: bool) -> None:
    """商談レコードに更新を適用する。"""
    con = sfa_db.connect(DB_PATH)
    try:
        hit = 0
        miss = 0
        for theme_id, fields in updates:
            deal = con.execute(
                "SELECT id, deal_name FROM deals WHERE theme_id=?", (theme_id,)
            ).fetchone()

            if not deal:
                print(f"  [SKIP]  theme_id={theme_id} — deals に対応レコードなし")
                miss += 1
                continue

            deal_id = deal["id"]
            deal_name = deal["deal_name"] or "(名称未設定)"
            sets = ", ".join(f"{col}=?" for col in fields) + ", updated_at=datetime('now')"
            params = list(fields.values()) + [deal_id]

            if dry_run:
                fields_str = ", ".join(f"{k}={v!r}" for k, v in fields.items())
                print(f"  [DRY]   theme_id={theme_id} deal_id={deal_id} {deal_name!r}")
                print(f"          → {fields_str}")
            else:
                con.execute(f"UPDATE deals SET {sets} WHERE id=?", params)
                print(f"  [OK]    theme_id={theme_id} deal_id={deal_id} {deal_name!r} "
                      f"({len(fields)}列更新)")
            hit += 1

        if not dry_run:
            con.commit()

        print()
        print(f"{'[DRY-RUN] ' if dry_run else ''}完了: 対象={hit}件 / 未対応={miss}件")
    finally:
        con.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="xlsxからコスト削減フィールドを商談DBに流し込む")
    parser.add_argument("--dry-run", action="store_true", help="書き込まずに確認だけ行う")
    args = parser.parse_args()

    print(f"xlsx読み込み: {XLSX_PATH}")
    rows = read_xlsx()
    print(f"  Salesレコード: {len(rows)}行")

    updates = build_updates(rows)
    print(f"  更新対象: {len(updates)}件（フィールドが1つ以上ある行）")
    print()

    if not updates:
        print("更新対象がありません。終了します。")
        return

    if args.dry_run:
        print("=== DRY-RUN モード（DBは変更されません）===")
    else:
        print(f"DB: {DB_PATH}")

    apply_updates(updates, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
