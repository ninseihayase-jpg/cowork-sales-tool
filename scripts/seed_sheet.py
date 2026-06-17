"""Google Sheet 初期投入用CSVを生成する（フェーズ1のオンボーディング）。

現行の sales_themes_input.xlsx「テーマKPI」を読み、運用イメージに沿って
  - Sales分（分類=Sales）            → seed_Sales.csv      … 営業メンバー全員が編集
  - Sales以外（分類≠Sales）          → seed_Sales以外.csv  … 主のみ編集
の2ファイルに分割出力する。各CSVをGoogle Sheetの対応タブに貼り付ければ初期データが揃う。

ヘッダ（列定義・プレフィックス）は元xlsxのまま保持するため、同期スクリプトがそのまま読める。

使い方:
  python scripts/seed_sheet.py [xlsx_path] [out_dir]
  （省略時: ../秘書/sales_themes_input.xlsx → ./seed/）
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT.parent.parent.parent.parent / "OneDrive" / "個人用" / "ツール作成" / "秘書" / "sales_themes_input.xlsx"


def _norm(h) -> str:
    return "" if h is None else str(h).split("\n")[-1].strip()


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "seed"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["テーマKPI"] if "テーマKPI" in wb.sheetnames else wb[wb.sheetnames[0]]

    raw_headers = [c.value for c in ws[1]]          # プレフィックス込みの元ヘッダ（そのまま保持）
    norm_headers = [_norm(h) for h in raw_headers]  # 分類判定用
    cat_idx = norm_headers.index("分類") if "分類" in norm_headers else None

    sales_rows, other_rows = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        category = (str(row[cat_idx]).strip() if cat_idx is not None and row[cat_idx] is not None else "")
        (sales_rows if category == "Sales" else other_rows).append(row)

    for fname, rows in [("seed_Sales.csv", sales_rows), ("seed_Sales以外.csv", other_rows)]:
        path = out_dir / fname
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(raw_headers)
            for r in rows:
                w.writerow(["" if v is None else v for v in r])
        print(f"出力: {path}  ({len(rows)}行)")

    print(f"\n完了。{out_dir} の2ファイルをGoogle Sheetの「Sales」「Sales以外」タブにそれぞれ貼り付けてください。")


if __name__ == "__main__":
    main()
