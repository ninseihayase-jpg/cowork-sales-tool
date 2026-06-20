"""xlsxから面談日を復元してテーマDB・営業情報DBに書き戻す。"""
from __future__ import annotations

import json
import os
import sqlite3
import sys
import urllib.request
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
for line in (ROOT / ".env").read_text().splitlines():
    if "=" in line and not line.startswith("#"):
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip())

sys.path.insert(0, str(ROOT))
from cowork import sfa_db

XLSX_PATH = Path("/mnt/c/Users/ninse/OneDrive/個人用/ツール作成/秘書/sales_themes_input.xlsx")
TOKEN = os.environ.get("THEME_API_TOKEN", "")
BASE_URL = os.environ.get("THEME_API_URL", "https://hisho-ohxe.onrender.com")
DB_PATH = os.environ.get("COWORK_SFA_DB", sfa_db.DEFAULT_DB_PATH)


def excel_to_iso(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    try:
        n = int(v)
        if n > 1000:
            return (date(1899, 12, 30) + timedelta(days=n)).isoformat()
    except (TypeError, ValueError):
        pass
    return None


def read_meeting_dates() -> dict[int, list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["テーマKPI"]
    headers = [str(c.value or "").split("\n")[-1].strip() for c in ws[1]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        r = dict(zip(headers, row))
        if r.get("分類") != "Sales":
            continue
        theme_id = int(r["ID"])
        dates = [d for d in (excel_to_iso(r.get(f"面談{i}")) for i in range(1, 10)) if d]
        if dates:
            result[theme_id] = dates
    return result


def restore_to_theme_db(meeting_data: dict[int, list[str]]) -> int:
    updated = 0
    for theme_id, dates in meeting_data.items():
        body = json.dumps({"sql": "UPDATE todos SET meeting_dates=? WHERE id=?",
                           "params": [json.dumps(dates), theme_id]}).encode()
        req = urllib.request.Request(
            f"{BASE_URL}/api/execute?token={TOKEN}", data=body,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            json.load(r)
        updated += 1
    return updated


def restore_to_sfa_db(meeting_data: dict[int, list[str]]) -> int:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    inserted = 0
    for theme_id, dates in meeting_data.items():
        deal = con.execute("SELECT id FROM deals WHERE theme_id=?", (theme_id,)).fetchone()
        if not deal:
            continue
        deal_id = deal["id"]
        for d in dates:
            exists = con.execute(
                "SELECT id FROM activities WHERE deal_id=? AND occurred_on=? AND type=?",
                (deal_id, d, "面談"),
            ).fetchone()
            if not exists:
                con.execute(
                    "INSERT INTO activities (deal_id, type, occurred_on, body) VALUES (?,?,?,?)",
                    (deal_id, "面談", d, "xlsxより復元"),
                )
                inserted += 1
    con.commit()
    con.close()
    return inserted


if __name__ == "__main__":
    meeting_data = read_meeting_dates()
    print(f"xlsx読み込み: {len(meeting_data)}件に面談日あり")

    print("テーマDBに書き戻し中...")
    n = restore_to_theme_db(meeting_data)
    print(f"  テーマDB更新: {n}件")

    print("営業情報DBに書き戻し中...")
    n = restore_to_sfa_db(meeting_data)
    print(f"  activities追加: {n}件")

    print("完了")
