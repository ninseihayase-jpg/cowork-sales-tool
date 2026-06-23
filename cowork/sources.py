"""入力データソース。ローカルxlsx と Google Sheets を同じインターフェイスで読む。

どちらも「ヘッダ正規化済みの dict のリスト」を返す。
これにより同期ロジック（sync.py）はソースの違いを意識しない。
"""

from __future__ import annotations

from pathlib import Path

from .mapping import normalize_header


def read_xlsx(path: str | Path, sheet_name: str = "テーマKPI") -> list[dict]:
    """ローカルxlsxを読む（フェーズ1のテスト/フォールバック用）。"""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [normalize_header(c.value) for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # ID列空はスキップ
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def _min_unused_id(used: set[int]) -> int:
    """used に含まれない最小の正整数を返す。"""
    i = 1
    while i in used:
        i += 1
    return i


def read_google_sheets(
    sheet_id: str,
    worksheet_names: list[str],
    service_account_json: str | Path,
    *,
    assign_ids: bool = False,
    existing_ids: set[int] | None = None,
) -> list[dict]:
    """Google Sheets を読む（フェーズ1本番）。複数ワークシートを連結して返す。

    Sales / Sales以外 のように分割したシートをまとめて1つのテーマ集合として扱う。
    各ワークシートは1行目がヘッダ（テーマKPIと同じ列定義）であること。

    assign_ids=True のとき:
      - ID列（先頭）が空の行に「空いている最小のID」を採番しスプシに書き戻す
      - 書き込み権限が必要（スコープ: spreadsheets）
      - existing_ids: テーマDB に既に存在するIDセット（重複を避けるために使用）
    """
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = (
        ["https://www.googleapis.com/auth/spreadsheets"]
        if assign_ids
        else ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    creds = Credentials.from_service_account_file(str(service_account_json), scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    rows: list[dict] = []
    for name in worksheet_names:
        try:
            ws = sh.worksheet(name)
        except gspread.WorksheetNotFound:
            continue
        values = ws.get_all_values(value_render_option="UNFORMATTED_VALUE")
        if not values:
            continue
        headers = [normalize_header(h) for h in values[0]]

        # スプシ内で使用済みのIDを収集（採番の重複防止）
        used_ids: set[int] = set(existing_ids or set())
        for raw in values[1:]:
            if raw and str(raw[0]).strip():
                try:
                    used_ids.add(int(str(raw[0]).strip()))
                except ValueError:
                    pass

        for row_idx, raw in enumerate(values[1:], start=2):  # 2行目から（1行目はヘッダ）
            padded = list(raw) + [""] * (len(headers) - len(raw))
            id_val = str(padded[0]).strip() if padded else ""

            if not id_val:
                if not assign_ids:
                    continue  # 採番しない場合はスキップ
                # 最小未使用IDを採番してスプシに書き戻す
                new_id = _min_unused_id(used_ids)
                used_ids.add(new_id)
                padded[0] = new_id
                ws.update_cell(row_idx, 1, new_id)
                print(f"[sources] {name} 行{row_idx}: ID採番 → {new_id}")

            rows.append(dict(zip(headers, padded)))
    return rows
