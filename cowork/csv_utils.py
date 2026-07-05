"""CSV取込共通ユーティリティ。BOM除去・改行コード正規化。"""

from __future__ import annotations


def normalize_csv_text(text: str) -> str:
    """CSVテキストの表記揺れを正規化する。

    - 先頭のUTF-8 BOM(﻿)を除去（Excel等のUTF-8 BOM付き出力をそのまま
      貼り付けると、ヘッダ列名の照合に失敗して全行が無言でスキップされるため）
    - 改行コードを\\nに統一する（\\r\\n・単独\\r混在時、csv.DictReaderが
      `new-line character seen in unquoted field` で例外を送出することがあるため）
    """
    if text.startswith("﻿"):
        text = text[1:]
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def xlsx_to_csv_text(data: bytes) -> str:
    """xlsxバイト列を先頭シートのCSVテキストに変換する。

    openpyxlのセル座標ベースのAPI（iter_rows）を使うため、空欄セルがあっても
    列がズレない（手組みのXMLパーサでは空欄セルの扱いを誤ると列がズレる）。
    """
    import csv
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if v is None else v for v in row])
    return buf.getvalue()
